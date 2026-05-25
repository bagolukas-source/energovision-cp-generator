"""ExcelReader — generický čítač .xlsx súborov s typovou bezpečnosťou.

Špecializované metódy pre:
    - eDistribúcia CSV/Excel formáty (SSE, ZSD, VSD)
    - Master Excel simulátor (Prepočet, Akumulácia sheets)
    - Faktúry distribútorov (extrakcia ročnej spotreby)
"""
from __future__ import annotations

from pathlib import Path
from typing import Literal, Optional

import pandas as pd
from openpyxl import load_workbook

from energovision_analytics.core.exceptions import DataIngestionError
from energovision_analytics.core.time_series import TimeSeriesData


class ExcelReader:
    """Generický čítač Excel s typovými helpermi."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise DataIngestionError(f"Excel súbor neexistuje: {self.path}")
        if self.path.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
            raise DataIngestionError(f"Nepodporovaný formát: {self.path.suffix}")

    # ------------------------------------------------------------------ Generic
    def list_sheets(self) -> list[str]:
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()

    def read_sheet(
        self,
        sheet_name: str,
        header_row: int = 0,
        usecols: Optional[str | list[str]] = None,
        skiprows: int = 0,
        nrows: Optional[int] = None,
    ) -> pd.DataFrame:
        """Načítaj konkrétny sheet do DataFrame."""
        try:
            df = pd.read_excel(
                self.path,
                sheet_name=sheet_name,
                header=header_row,
                usecols=usecols,
                skiprows=skiprows,
                nrows=nrows,
                engine="openpyxl",
            )
        except Exception as e:
            raise DataIngestionError(f"Chyba pri čítaní {sheet_name}: {e}") from e
        return df

    def read_range(
        self,
        sheet_name: str,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> list[list]:
        """Načítaj konkrétny range cells (1-indexed, vrátane oboch koncov)."""
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(
                min_row=start_row, max_row=end_row,
                min_col=start_col, max_col=end_col,
                values_only=True,
            ):
                data.append(list(row))
            return data
        finally:
            wb.close()

    # ------------------------------------------------------------------ eDistribúcia CSV
    @staticmethod
    def read_edistribucia_csv(
        path: str | Path,
        distribuutor: Literal["SSE", "ZSD", "VSD"],
        granularity_min: Literal[15, 60] = 15,
    ) -> TimeSeriesData:
        """Načítaj 15-min profil z CSV stiahnutého z eDistribúcia portálu.

        Známe formáty (k 2026):
            SSE: stĺpce 'DateTime', 'Active+ [kWh]' (15 min)
            ZSD: stĺpce 'Datum', 'Cas', 'CinneSpotrebovane' (15 min)
            VSD: stĺpce 'Časová značka', 'Spotreba [kWh]' (15 min)

        Pretože nemáme vzorky lokálne, používame **tolerantný parser** s heuristickou
        detekciou stĺpcov. Pri reálnom nasadení sa špecifikuje per distribútor.

        Args:
            path: Cesta k CSV súboru
            distribuutor: 'SSE' | 'ZSD' | 'VSD'
            granularity_min: Najčastejšie 15

        Returns:
            TimeSeriesData s kW hodnotami (konverzia kWh × (60/granularity_min))
        """
        path = Path(path)
        if not path.exists():
            raise DataIngestionError(f"CSV neexistuje: {path}")

        # Načítaj tolerantne — skús viacero oddelovačov
        for sep in [";", ",", "\t"]:
            try:
                df = pd.read_csv(path, sep=sep, encoding="utf-8")
                if len(df.columns) >= 2:
                    break
            except Exception:
                continue
        else:
            raise DataIngestionError(f"Nedalo sa parsovať CSV {path}")

        # Heuristika — nájdi timestamp stĺpec
        ts_candidates = [
            c for c in df.columns
            if any(k in c.lower() for k in ("datetime", "datum", "čas", "cas", "time", "timestamp"))
        ]
        if not ts_candidates:
            raise DataIngestionError(
                f"Nenašiel som timestamp stĺpec v {path}. Stĺpce: {list(df.columns)}"
            )
        ts_col = ts_candidates[0]

        # Heuristika — nájdi value stĺpec (kWh alebo kW)
        val_candidates = [
            c for c in df.columns
            if any(k in c.lower() for k in ("kwh", "kw", "spotreba", "active", "energy"))
        ]
        if not val_candidates:
            raise DataIngestionError(
                f"Nenašiel som value stĺpec v {path}. Stĺpce: {list(df.columns)}"
            )
        val_col = val_candidates[0]

        # Parse timestampy
        df[ts_col] = pd.to_datetime(df[ts_col], dayfirst=True, errors="coerce")
        df = df.dropna(subset=[ts_col, val_col])
        df = df.sort_values(ts_col).reset_index(drop=True)

        # Konverzia kWh → kW (kWh × intervals_per_hour)
        intervals_per_hour = 60 / granularity_min
        is_kwh = "kwh" in val_col.lower() or "spotreba" in val_col.lower() or "energy" in val_col.lower()
        values_kw = (
            df[val_col].astype(float) * intervals_per_hour
            if is_kwh else df[val_col].astype(float)
        ).tolist()

        return TimeSeriesData(
            timestamps=df[ts_col].tolist(),
            values=values_kw,
            granularity_min=granularity_min,
            column_name="load_kw",
            unit="kW",
            source=f"eDistribucia_{distribuutor}",
        )

    # ------------------------------------------------------------------ Master Excel (Energovision)
    def read_master_prepocet(self) -> dict[str, list[float]]:
        """Načítaj Prepočet sheet z master Excelu (consumption + PV per hodina).

        Master Excel formát:
            Sheet "Prepočet", stĺpce D (consumption kW), E (PV kW)
            Riadky 11-8770 (8760 hodín)
        """
        sheets = self.list_sheets()
        if "Prepočet" not in sheets:
            raise DataIngestionError(
                f"Sheet 'Prepočet' chýba v {self.path.name}. Dostupné: {sheets}"
            )

        data = self.read_range("Prepočet", start_row=11, end_row=8770,
                                start_col=4, end_col=5)

        consumption_kw = [float(row[0]) if row[0] is not None else 0.0 for row in data]
        pv_kw = [float(row[1]) if row[1] is not None else 0.0 for row in data]

        return {
            "consumption_kw": consumption_kw,
            "pv_kw": pv_kw,
        }

    def read_master_spot_prices(self) -> list[Optional[float]]:
        """Načítaj spot ceny z Akumulácia sheet stĺpec P."""
        sheets = self.list_sheets()
        if "Akumulácia" not in sheets:
            raise DataIngestionError(
                f"Sheet 'Akumulácia' chýba. Dostupné: {sheets}"
            )

        data = self.read_range("Akumulácia", start_row=12, end_row=8771,
                                start_col=16, end_col=16)

        return [float(row[0]) if row[0] is not None else None for row in data]

    def read_master_summary(self) -> dict[str, float]:
        """Načítaj ročné sumáre zo sheet Prepočet (pre validation/regression)."""
        sheets = self.list_sheets()
        if "Prepočet" not in sheets:
            raise DataIngestionError(f"Sheet 'Prepočet' chýba. Dostupné: {sheets}")

        summary: dict[str, float] = {}
        for label, (col, row) in [
            ("load_total_kwh", (4, 8)),
            ("pv_total_kwh", (5, 8)),
            ("pv_direct_kwh", (6, 8)),
            ("pv_to_grid_kwh", (7, 8)),
        ]:
            cell_data = self.read_range("Prepočet", row, row, col, col)
            if cell_data and cell_data[0][0] is not None:
                summary[label] = float(cell_data[0][0])

        return summary
